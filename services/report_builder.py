import re
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_COLUMNS = (
    ("brand", "Бренд"),
    ("article", "Артикул"),
    ("name", "Наименование"),
    ("manufacturer", "Производитель"),
    ("country_origin", "Страна происхождения"),
    ("quantity", "Количество"),
    ("unit", "Единица измерения"),
    ("tnved_code", "Код ТН ВЭД"),
    ("permit_document", "Разрешительная документация"),
    ("permit_number", "Номер разрешительного документа"),
    ("permit_expiry", "Срок действия документа"),
    ("comment", "Комментарий"),
)

RESULT_COLUMNS = (
    "Статус",
    "Уровень риска",
    "Замечание",
    "Что проверить",
)


def build_excel_report(
    source_filename: str,
    records: list[dict[str, str]],
    warnings: list[str],
    ai_analysis: str,
) -> bytes:
    """Формирует Excel-отчёт и возвращает его содержимое в байтах."""

    workbook = Workbook()

    check_sheet = workbook.active
    check_sheet.title = "Проверка"

    _fill_check_sheet(
        worksheet=check_sheet,
        records=records,
    )

    summary_sheet = workbook.create_sheet("Итоги")

    _fill_summary_sheet(
        worksheet=summary_sheet,
        source_filename=source_filename,
        records_count=len(records),
        warnings=warnings,
    )

    ai_sheet = workbook.create_sheet("AI-анализ")

    _fill_ai_sheet(
        worksheet=ai_sheet,
        ai_analysis=ai_analysis,
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue()


def build_report_filename(source_filename: str) -> str:
    """Создаёт безопасное название итогового отчёта."""

    source_stem = Path(source_filename).stem

    safe_stem = re.sub(
        r'[<>:"/\\|?*]+',
        "_",
        source_stem,
    ).strip()

    if not safe_stem:
        safe_stem = "file"

    return f"{safe_stem}_VED_AI_report.xlsx"


def _fill_check_sheet(
    worksheet,
    records: list[dict[str, str]],
) -> None:
    headers = [
        header
        for _, header in SOURCE_COLUMNS
    ]

    headers.extend(RESULT_COLUMNS)

    worksheet.append(headers)

    _style_header(worksheet)

    for record in records:
        evaluation = _evaluate_record(record)

        row = [
            record.get(field_name, "")
            for field_name, _ in SOURCE_COLUMNS
        ]

        row.extend(
            [
                evaluation["status"],
                evaluation["risk"],
                evaluation["remark"],
                evaluation["action"],
            ]
        )

        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    _format_check_sheet(worksheet)


def _evaluate_record(
    record: dict[str, str],
) -> dict[str, str]:
    problems: list[str] = []
    actions: list[str] = []

    name = record.get("name", "").strip()
    article = record.get("article", "").strip()
    brand = record.get("brand", "").strip()
    quantity = record.get("quantity", "").strip()
    tnved_code = record.get("tnved_code", "").strip()
    permit = record.get("permit_document", "").strip()

    if not name:
        problems.append(
            "Не указано полное наименование товара."
        )
        actions.append(
            "Добавить коммерческое и техническое наименование."
        )

    if not article:
        problems.append(
            "Не указан артикул или каталожный номер."
        )
        actions.append(
            "Уточнить артикул, модель или каталожный номер."
        )

    if not brand:
        problems.append(
            "Не указан бренд."
        )
        actions.append(
            "Уточнить бренд или изготовителя."
        )

    if not quantity:
        problems.append(
            "Не указано количество."
        )
        actions.append(
            "Добавить количество и единицу измерения."
        )
    elif not _is_positive_number(quantity):
        problems.append(
            "Количество имеет некорректное значение."
        )
        actions.append(
            "Проверить, что количество является положительным числом."
        )

    if not tnved_code:
        problems.append(
            "Код ТН ВЭД отсутствует."
        )
        actions.append(
            "Определить код по техническим характеристикам товара."
        )
    else:
        code_digits = re.sub(
            r"\D",
            "",
            tnved_code,
        )

        if len(code_digits) != 10:
            problems.append(
                "Заявленный код ТН ВЭД состоит не из 10 цифр."
            )
            actions.append(
                "Исправить формат кода ТН ВЭД."
            )
        else:
            actions.append(
                "Проверить заявленный код по описанию, назначению, "
                "материалу и техническим характеристикам."
            )

    if not permit:
        problems.append(
            "В файле нет сведений о разрешительной документации."
        )
        actions.append(
            "Определить, требуется ли подтверждение соответствия, "
            "лицензия, маркировка или другой документ."
        )

    if problems:
        status = "Требует проверки"
        risk = "Жёлтый"
        remark = " ".join(problems)
    else:
        status = "Предварительно заполнено"
        risk = "Зелёный"
        remark = (
            "Основные поля заполнены. Код ТН ВЭД и разрешительная "
            "документация заявлены, но системой не подтверждены."
        )

    return {
        "status": status,
        "risk": risk,
        "remark": remark,
        "action": " ".join(dict.fromkeys(actions)),
    }


def _is_positive_number(value: str) -> bool:
    normalized = (
        value
        .replace(" ", "")
        .replace(",", ".")
    )

    try:
        return float(normalized) > 0
    except ValueError:
        return False


def _fill_summary_sheet(
    worksheet,
    source_filename: str,
    records_count: int,
    warnings: list[str],
) -> None:
    worksheet["A1"] = "Отчёт VED AI"
    worksheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    worksheet["A3"] = "Исходный файл"
    worksheet["B3"] = source_filename

    worksheet["A4"] = "Количество позиций"
    worksheet["B4"] = records_count

    worksheet["A6"] = "Общие замечания"
    worksheet["A6"].font = Font(bold=True)

    if warnings:
        for row_number, warning in enumerate(
            warnings,
            start=7,
        ):
            worksheet.cell(
                row=row_number,
                column=1,
                value=f"• {warning}",
            )
    else:
        worksheet["A7"] = "Автоматические замечания отсутствуют."

    worksheet["A10"] = (
        "Коды ТН ВЭД и разрешительные документы в отчёте "
        "считаются заявленными пользователем, а не подтверждёнными."
    )

    worksheet["A10"].alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    worksheet.column_dimensions["A"].width = 75
    worksheet.column_dimensions["B"].width = 40


def _fill_ai_sheet(
    worksheet,
    ai_analysis: str,
) -> None:
    worksheet["A1"] = "Предварительный AI-анализ"
    worksheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    worksheet["A3"] = (
        ai_analysis
        or "AI-анализ не был сформирован."
    )

    worksheet["A3"].alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    worksheet.column_dimensions["A"].width = 120
    worksheet.row_dimensions[3].height = 400


def _style_header(worksheet) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _format_check_sheet(worksheet) -> None:
    widths = {
        "A": 18,
        "B": 20,
        "C": 35,
        "D": 25,
        "E": 22,
        "F": 14,
        "G": 16,
        "H": 18,
        "I": 35,
        "J": 25,
        "K": 20,
        "L": 25,
        "M": 22,
        "N": 18,
        "O": 55,
        "P": 60,
    }

    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    for row in worksheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        if column_letter not in widths:
            worksheet.column_dimensions[
                column_letter
            ].width = 20