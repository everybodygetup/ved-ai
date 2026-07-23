import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from itertools import chain, islice
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pypdf import PdfReader
from pypdf.errors import PdfReadError


MAX_EXCEL_ROWS = 5000
MAX_PDF_PAGES = 30
PREVIEW_ROWS_COUNT = 5


class FileParserError(Exception):
    """Понятная пользователю ошибка разбора файла."""


@dataclass(slots=True)
class ParsedFile:
    filename: str
    file_type: str

    sheet_name: str | None = None
    page_count: int | None = None

    mapped_columns: dict[str, str] = field(default_factory=dict)
    records: list[dict[str, str]] = field(default_factory=list)

    total_rows: int = 0
    text_preview: str = ""

    warnings: list[str] = field(default_factory=list)


FIELD_LABELS = {
    "brand": "Бренд",
    "article": "Артикул",
    "name": "Наименование",
    "manufacturer": "Производитель",
    "country_origin": "Страна происхождения",
    "quantity": "Количество",
    "unit": "Единица измерения",
    "tnved_code": "Код ТН ВЭД",
    "permit_document": "Разрешительная документация",
    "permit_number": "Номер разрешительного документа",
    "permit_expiry": "Срок действия документа",
    "comment": "Комментарий",
}


HEADER_ALIASES = {
    "brand": {
        "бренд",
        "brand",
        "марка",
        "товарный знак",
    },
    "article": {
        "артикул",
        "article",
        "part number",
        "part no",
        "номер детали",
        "каталожный номер",
        "код товара",
        "sku",
    },
    "name": {
        "наименование",
        "наименование товара",
        "описание",
        "описание товара",
        "товар",
        "product",
        "product name",
        "description",
    },
    "manufacturer": {
        "производитель",
        "изготовитель",
        "manufacturer",
        "producer",
    },
    "country_origin": {
        "страна происхождения",
        "страна производства",
        "country of origin",
        "origin country",
        "origin",
    },
    "quantity": {
        "количество",
        "кол во",
        "количество товара",
        "qty",
        "quantity",
    },
    "unit": {
        "единица измерения",
        "ед изм",
        "unit",
        "unit of measure",
        "uom",
    },
    "tnved_code": {
        "код тн вэд",
        "тн вэд",
        "код тн вэд еаэс",
        "код тнвэд",
        "tn ved",
        "tnved",
        "hs code",
        "customs code",
    },
    "permit_document": {
        "разрешительная документация",
        "разрешительный документ",
        "сертификат декларация",
        "сертификат",
        "декларация соответствия",
        "certificate",
        "permit document",
    },
    "permit_number": {
        "номер разрешительного документа",
        "номер сертификата",
        "номер декларации",
        "certificate number",
        "document number",
    },
    "permit_expiry": {
        "срок действия",
        "действителен до",
        "дата окончания",
        "expiry date",
        "valid until",
    },
    "comment": {
        "комментарий",
        "примечание",
        "comment",
        "note",
    },
}


def parse_uploaded_file(
    filename: str,
    file_object: BinaryIO,
) -> ParsedFile:
    """Определяет формат файла и запускает нужный обработчик."""

    extension = Path(filename).suffix.lower()

    file_object.seek(0)

    if extension == ".xlsx":
        return _parse_excel(filename, file_object)

    if extension == ".pdf":
        return _parse_pdf(filename, file_object)

    raise FileParserError(
        "Поддерживаются только файлы .xlsx и .pdf."
    )


def format_parse_result(result: ParsedFile) -> str:
    """Формирует удобный отчёт для Telegram."""

    if result.file_type == "excel":
        return _format_excel_result(result)

    return _format_pdf_result(result)


def _parse_excel(
    filename: str,
    file_object: BinaryIO,
) -> ParsedFile:
    try:
        workbook = load_workbook(
            filename=file_object,
            read_only=True,
            data_only=True,
        )
    except (
        InvalidFileException,
        OSError,
        ValueError,
        KeyError,
    ) as error:
        raise FileParserError(
            "Не удалось прочитать Excel. "
            "Проверьте, что это исправный файл .xlsx."
        ) from error

    try:
        if not workbook.sheetnames:
            raise FileParserError(
                "В Excel-файле не найдено ни одного листа."
            )

        worksheet = workbook[workbook.sheetnames[0]]

        rows_iterator = worksheet.iter_rows(values_only=True)

        sample_rows = list(
            islice(rows_iterator, 25)
        )

        if not sample_rows:
            raise FileParserError(
                "Excel-файл не содержит данных."
            )

        header_index, mapped_indices = _detect_header_row(
            sample_rows
        )

        header_row = sample_rows[header_index]

        original_headers = [
            _stringify(value)
            for value in header_row
        ]

        mapped_columns = {
            field_name: original_headers[column_index]
            for field_name, column_index in mapped_indices.items()
        }

        result = ParsedFile(
            filename=filename,
            file_type="excel",
            sheet_name=worksheet.title,
            mapped_columns=mapped_columns,
        )

        if len(mapped_indices) < 2:
            result.warnings.append(
                "Структура таблицы распознана частично. "
                "Проверьте названия колонок."
            )

        data_rows = chain(
            sample_rows[header_index + 1 :],
            rows_iterator,
        )

        for row in data_rows:
            if result.total_rows >= MAX_EXCEL_ROWS:
                result.warnings.append(
                    f"Обработаны только первые "
                    f"{MAX_EXCEL_ROWS} позиций."
                )
                break

            if not any(
                _stringify(value)
                for value in row
            ):
                continue

            record = {}

            for field_name, column_index in mapped_indices.items():
                value = (
                    row[column_index]
                    if column_index < len(row)
                    else None
                )

                record[field_name] = _stringify(value)

            if not any(record.values()):
                continue

            result.records.append(record)
            result.total_rows += 1

        if not result.records:
            raise FileParserError(
                "После строки заголовков позиции не найдены."
            )

        _validate_excel(result)

        return result

    finally:
        workbook.close()


def _parse_pdf(
    filename: str,
    file_object: BinaryIO,
) -> ParsedFile:
    try:
        reader = PdfReader(file_object)
    except (
        PdfReadError,
        OSError,
        ValueError,
    ) as error:
        raise FileParserError(
            "Не удалось прочитать PDF-файл."
        ) from error

    if reader.is_encrypted:
        try:
            reader.decrypt("")
        except Exception as error:
            raise FileParserError(
                "PDF защищён паролем. "
                "Загрузите файл без защиты."
            ) from error

    page_count = len(reader.pages)

    extracted_parts = []

    for page in reader.pages[:MAX_PDF_PAGES]:
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""

        text = text.strip()

        if text:
            extracted_parts.append(text)

    full_text = "\n\n".join(extracted_parts).strip()

    if not full_text:
        raise FileParserError(
            "В PDF не найден текст. Вероятно, это скан. "
            "Для таких файлов позже добавим OCR."
        )

    result = ParsedFile(
        filename=filename,
        file_type="pdf",
        page_count=page_count,
        text_preview=full_text[:3000],
    )

    if page_count > MAX_PDF_PAGES:
        result.warnings.append(
            f"Извлечены только первые {MAX_PDF_PAGES} страниц."
        )

    result.warnings.append(
        "Из PDF извлечён текст, но структура таблицы "
        "пока не гарантируется."
    )

    return result


def _detect_header_row(
    rows: list[tuple],
) -> tuple[int, dict[str, int]]:
    best_index = 0
    best_mapping: dict[str, int] = {}
    best_score = -1

    for row_index, row in enumerate(rows):
        current_mapping: dict[str, int] = {}

        for column_index, value in enumerate(row):
            field_name = _recognize_header(value)

            if (
                field_name
                and field_name not in current_mapping
            ):
                current_mapping[field_name] = column_index

        score = len(current_mapping)

        if score > best_score:
            best_score = score
            best_index = row_index
            best_mapping = current_mapping

    return best_index, best_mapping


def _recognize_header(value: object) -> str | None:
    normalized = _normalize_header(value)

    if not normalized:
        return None

    for field_name, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return field_name

    for field_name, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if len(alias) >= 5 and alias in normalized:
                return field_name

    return None


def _normalize_header(value: object) -> str:
    text = _stringify(value).lower().replace("ё", "е")

    text = re.sub(
        r"[^a-zа-я0-9]+",
        " ",
        text,
        flags=re.IGNORECASE,
    )

    return re.sub(r"\s+", " ", text).strip()


def _stringify(value: object) -> str:
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def _validate_excel(result: ParsedFile) -> None:
    recommended_fields = {
        "brand",
        "article",
        "name",
        "quantity",
        "tnved_code",
        "permit_document",
    }

    missing_columns = [
        FIELD_LABELS[field]
        for field in recommended_fields
        if field not in result.mapped_columns
    ]

    if missing_columns:
        result.warnings.append(
            "Не найдены рекомендуемые колонки: "
            + ", ".join(sorted(missing_columns))
            + "."
        )

    invalid_codes = 0
    empty_codes = 0
    empty_permits = 0
    invalid_quantities = 0

    article_codes: dict[str, set[str]] = {}

    for record in result.records:
        code = record.get("tnved_code", "")
        article = record.get("article", "")
        permit = record.get("permit_document", "")
        quantity = record.get("quantity", "")

        if "tnved_code" in result.mapped_columns:
            if not code:
                empty_codes += 1
            else:
                digits = re.sub(r"\D", "", code)

                if len(digits) != 10:
                    invalid_codes += 1

                if article:
                    article_codes.setdefault(
                        article.lower(),
                        set(),
                    ).add(digits)

        if (
            "permit_document" in result.mapped_columns
            and not permit
        ):
            empty_permits += 1

        if (
            "quantity" in result.mapped_columns
            and quantity
            and not _is_positive_number(quantity)
        ):
            invalid_quantities += 1

    conflicting_articles = sum(
        1
        for codes in article_codes.values()
        if len(codes) > 1
    )

    if empty_codes:
        result.warnings.append(
            f"Позиций без кода ТН ВЭД: {empty_codes}."
        )

    if invalid_codes:
        result.warnings.append(
            f"Позиций с кодом не из 10 цифр: {invalid_codes}."
        )

    if empty_permits:
        result.warnings.append(
            "Позиций без указанной разрешительной "
            f"документации: {empty_permits}."
        )

    if invalid_quantities:
        result.warnings.append(
            "Позиций с некорректным количеством: "
            f"{invalid_quantities}."
        )

    if conflicting_articles:
        result.warnings.append(
            "Артикулов, которым присвоены разные коды "
            f"ТН ВЭД: {conflicting_articles}."
        )


def _is_positive_number(value: str) -> bool:
    normalized = (
        value
        .replace(" ", "")
        .replace(",", ".")
    )

    try:
        return float(normalized) > 0
    except ValueError:
        return False


def _format_excel_result(result: ParsedFile) -> str:
    lines = [
        "✅ Excel-файл прочитан",
        "",
        f"📄 Файл: {result.filename}",
        f"📑 Лист: {result.sheet_name}",
        f"📦 Найдено позиций: {result.total_rows}",
        "",
        "🔎 Распознанные колонки:",
    ]

    if result.mapped_columns:
        for field_name, original_name in result.mapped_columns.items():
            lines.append(
                f"• {FIELD_LABELS[field_name]}: "
                f"«{original_name}»"
            )
    else:
        lines.append("• Колонки не распознаны")

    lines.extend(
        [
            "",
            "👀 Первые позиции:",
        ]
    )

    for row_number, record in enumerate(
        result.records[:PREVIEW_ROWS_COUNT],
        start=1,
    ):
        values = []

        for field_name in (
            "brand",
            "article",
            "name",
            "quantity",
            "tnved_code",
            "permit_document",
        ):
            value = record.get(field_name)

            if value:
                values.append(
                    f"{FIELD_LABELS[field_name]}: {value}"
                )

        row_text = "; ".join(values) or "данные не распознаны"

        lines.append(
            f"{row_number}. {row_text}"
        )

    if result.warnings:
        lines.extend(
            [
                "",
                "⚠️ Что требует внимания:",
            ]
        )

        for warning in result.warnings:
            lines.append(f"• {warning}")

    lines.extend(
        [
            "",
            "Коды ТН ВЭД и разрешительные документы "
            "сейчас считаются заявленными в файле, "
            "а не подтверждёнными системой.",
        ]
    )

    return "\n".join(lines)


def _format_pdf_result(result: ParsedFile) -> str:
    lines = [
        "✅ PDF-файл прочитан",
        "",
        f"📄 Файл: {result.filename}",
        f"📑 Страниц: {result.page_count}",
        "",
        "📝 Извлечённый текст:",
        "",
        result.text_preview,
    ]

    if result.warnings:
        lines.extend(
            [
                "",
                "⚠️ Примечания:",
            ]
        )

        for warning in result.warnings:
            lines.append(f"• {warning}")

    return "\n".join(lines)

def build_file_llm_request(
    result: ParsedFile,
    max_records: int = 30,
) -> str:
    """
    Создаёт обезличенную сводку для LLM.

    Полный исходный файл модели не передаётся.
    """

    if result.file_type == "pdf":
        return (
            "Проведи предварительный анализ текста документа.\n\n"
            f"Количество страниц: {result.page_count}\n\n"
            "Извлечённый текст:\n"
            f"{result.text_preview}\n\n"
            "Не придумывай нормы, коды и требования. "
            "Укажи, какие сведения необходимо проверить вручную."
        )

    lines = [
        "Проведи предварительный анализ товарного перечня.",
        "",
        f"Всего позиций в файле: {result.total_rows}.",
        "",
        "Заявленные позиции:",
    ]

    fields = (
        "brand",
        "article",
        "name",
        "manufacturer",
        "country_origin",
        "quantity",
        "unit",
        "tnved_code",
        "permit_document",
        "permit_number",
        "permit_expiry",
    )

    for number, record in enumerate(
        result.records[:max_records],
        start=1,
    ):
        values = []

        for field_name in fields:
            value = record.get(field_name)

            if value:
                values.append(
                    f"{FIELD_LABELS[field_name]}: {value}"
                )

        lines.append(
            f"{number}. " + "; ".join(values)
        )

    if result.total_rows > max_records:
        lines.append(
            f"\nПоказаны только первые {max_records} позиций."
        )

    if result.warnings:
        lines.extend(
            [
                "",
                "Результаты автоматической проверки:",
            ]
        )

        for warning in result.warnings:
            lines.append(f"- {warning}")

    lines.extend(
        [
            "",
            "Правила анализа:",
            "- коды ТН ВЭД считаются заявленными, а не проверенными;",
            "- разрешительные документы считаются заявленными;",
            "- не подтверждай правильность кода без технического описания;",
            "- найди отсутствующие и противоречивые сведения;",
            "- раздели риски на красные, жёлтые и зелёные;",
            "- предложи следующие действия;",
            "- ответ предназначен для Telegram, без таблиц и HTML.",
            "- отсутствие данных не считать подтверждённым нарушением;",
            "- различать подтверждённую проблему, потенциальный риск и недостаток данных;",
            "- не утверждать, что заявленный код неправильный, если нет технического описания;",
            "- не считать отсутствие сведений о разрешительном документе автоматическим запретом ввоза;",
            "- красный риск указывать только при наличии конкретного подтверждённого основания;",
        ]
    )

    return "\n".join(lines)